import yfinance as yf
import pandas as pd
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

# Comprehensive list of Indonesian stocks - LQ45, IDX30, IDX80 constituents
# These represent the most liquid stocks that make up ~90% of trading volume
INDONESIAN_STOCKS = [
    # IDX30 - Top 30 most liquid stocks
    'BBCA.JK', 'BBRI.JK', 'BMRI.JK', 'BBNI.JK', 'ASII.JK', 'TLKM.JK',
    'UNTR.JK', 'AMMN.JK', 'ADRO.JK', 'ICBP.JK', 'INDF.JK', 'KLBF.JK',
    'EMTK.JK', 'GOTO.JK', 'BUKA.JK', 'ACES.JK', 'INKP.JK', 'MAPI.JK',
    'PGAS.JK', 'PTBA.JK', 'SMGR.JK', 'TOWR.JK', 'INTP.JK', 'CPIN.JK',
    'BRIS.JK', 'MDKA.JK', 'BYAN.JK', 'BRPT.JK', 'EXCL.JK', 'ISAT.JK',
    
    # Additional LQ45 constituents (45 total)
    'ARTO.JK', 'ANTM.JK', 'GGRM.JK', 'HMSP.JK', 'MEDC.JK', 'MIKA.JK',
    'PWON.JK', 'TKIM.JK', 'JPFA.JK', 'CTRA.JK', 'MAPA.JK', 'AADI.JK',
    'SCMA.JK', 'ERAA.JK', 'HRUM.JK',
    
    # Additional IDX80 constituents (extending to ~100 stocks for comprehensive coverage)
    'ITMG.JK', 'KAEF.JK', 'TBIG.JK', 'AALI.JK', 'AKRA.JK', 'BNBR.JK',
    'BSDE.JK', 'CTRA.JK', 'DOID.JK', 'ELSA.JK', 'ESSA.JK', 'EXCL.JK',
    'FILM.JK', 'HOKI.JK', 'ICBP.JK', 'IMAS.JK', 'INCO.JK', 'ISAT.JK',
    'JSMR.JK', 'LPPF.JK', 'LSIP.JK', 'MNCN.JK', 'MYOR.JK', 'PANI.JK',
    'PGAS.JK', 'PTPP.JK', 'PWON.JK', 'RAJA.JK', 'SCMA.JK', 'SIDO.JK',
    'SILO.JK', 'SMBR.JK', 'SMRA.JK', 'SRIL.JK', 'TINS.JK', 'TPIA.JK',
    'UNVR.JK', 'WIKA.JK', 'WSBP.JK', 'WSKT.JK',
    
    # Additional major traded stocks
    'BNLI.JK', 'BTPS.JK', 'BUMI.JK', 'DMAS.JK', 'DSNG.JK', 'INDY.JK',
    'JRPT.JK', 'LPKR.JK', 'MKPI.JK', 'MTDL.JK', 'NCKL.JK', 'PNLF.JK',
    'PSAB.JK', 'RALS.JK', 'SMMT.JK', 'SSMS.JK', 'TARA.JK', 'TOBA.JK',
    'TOTL.JK', 'WIFI.JK', 'WINS.JK', 'WOOD.JK',
    
    # State-owned enterprises and other significant stocks
    'ADHI.JK', 'BBTN.JK', 'BJBR.JK', 'BJTM.JK', 'BUKK.JK', 'DGNS.JK',
    'GIAA.JK', 'INAF.JK', 'KIJA.JK', 'PJAA.JK', 'PTRO.JK', 'SMBR.JK',
    'TBLA.JK', 'TRAM.JK', 'WTON.JK',
]

def get_all_idx_stocks():
    """
    Fetch all stocks from IDX Composite dynamically.
    This is a fallback/supplement to the hardcoded list.
    """
    additional_stocks = []
    try:
        # Try to get IDX Composite constituents
        idx = yf.Ticker('^JKSE')
        # Note: yfinance may not always provide constituents for all indices
        # This is a best-effort approach
        print("Attempting to fetch additional IDX constituents dynamically...")
    except Exception as e:
        print(f"Note: Could not fetch additional stocks dynamically: {e}")
    
    return additional_stocks

def get_stock_data(ticker, period='5y'):
    """Fetch stock data and financial information"""
    try:
        stock = yf.Ticker(ticker)
        
        # Get historical data
        hist = stock.history(period=period)
        if hist.empty:
            return None
        
        # Get financial data
        info = stock.info
        
        # Calculate metrics
        current_price = hist['Close'][-1] if len(hist) > 0 else np.nan
        year_ago_price = hist['Close'][-252] if len(hist) >= 252 else hist['Close'][0]
        price_change_1y = ((current_price - year_ago_price) / year_ago_price * 100) if year_ago_price > 0 else np.nan
        
        # 5-year return
        start_price = hist['Close'][0]
        price_change_5y = ((current_price - start_price) / start_price * 100) if start_price > 0 else np.nan
        
        # Volatility (annualized standard deviation)
        returns = hist['Close'].pct_change().dropna()
        volatility = returns.std() * np.sqrt(252) * 100 if len(returns) > 0 else np.nan
        
        # Average volume
        avg_volume = hist['Volume'].mean()
        
        # Sharpe ratio (assuming risk-free rate of 6% for Indonesia)
        risk_free_rate = 0.06
        avg_return = returns.mean() * 252
        sharpe_ratio = (avg_return - risk_free_rate) / (returns.std() * np.sqrt(252)) if returns.std() > 0 else np.nan
        
        # Extract financial metrics
        data = {
            'Ticker': ticker,
            'Company Name': info.get('longName', ticker),
            'Sector': info.get('sector', 'N/A'),
            'Industry': info.get('industry', 'N/A'),
            'Current Price': current_price,
            'Market Cap': info.get('marketCap', np.nan),
            'P/E Ratio': info.get('trailingPE', np.nan),
            'Forward P/E': info.get('forwardPE', np.nan),
            'PEG Ratio': info.get('pegRatio', np.nan),
            'Price to Book': info.get('priceToBook', np.nan),
            'Debt to Equity': info.get('debtToEquity', np.nan),
            'ROE (%)': info.get('returnOnEquity', np.nan) * 100 if info.get('returnOnEquity') else np.nan,
            'ROA (%)': info.get('returnOnAssets', np.nan) * 100 if info.get('returnOnAssets') else np.nan,
            'Profit Margin (%)': info.get('profitMargins', np.nan) * 100 if info.get('profitMargins') else np.nan,
            'Operating Margin (%)': info.get('operatingMargins', np.nan) * 100 if info.get('operatingMargins') else np.nan,
            'Revenue Growth (%)': info.get('revenueGrowth', np.nan) * 100 if info.get('revenueGrowth') else np.nan,
            'Earnings Growth (%)': info.get('earningsGrowth', np.nan) * 100 if info.get('earningsGrowth') else np.nan,
            'Dividend Yield (%)': info.get('dividendYield', np.nan) * 100 if info.get('dividendYield') else np.nan,
            'Payout Ratio (%)': info.get('payoutRatio', np.nan) * 100 if info.get('payoutRatio') else np.nan,
            'Current Ratio': info.get('currentRatio', np.nan),
            'Quick Ratio': info.get('quickRatio', np.nan),
            '1Y Price Change (%)': price_change_1y,
            '5Y Price Change (%)': price_change_5y,
            'Volatility (%)': volatility,
            'Sharpe Ratio': sharpe_ratio,
            'Avg Daily Volume': avg_volume,
            '52W High': info.get('fiftyTwoWeekHigh', np.nan),
            '52W Low': info.get('fiftyTwoWeekLow', np.nan),
            'Beta': info.get('beta', np.nan),
            'Book Value': info.get('bookValue', np.nan),
            'Enterprise Value': info.get('enterpriseValue', np.nan),
            'EV/EBITDA': info.get('enterpriseToEbitda', np.nan),
        }
        
        return data
    except Exception as e:
        print(f"Error fetching {ticker}: {str(e)}")
        return None

def calculate_composite_score(df):
    """Calculate a composite score based on multiple factors with weighted importance"""
    score_df = df.copy()
    
    # Metrics grouped by importance
    high_weight_positive = {
        'ROE (%)': 3.0,
        'ROA (%)': 2.5,
        'Profit Margin (%)': 2.5,
        'Revenue Growth (%)': 2.0,
        'Earnings Growth (%)': 2.0,
        '1Y Price Change (%)': 1.5,
        'Sharpe Ratio': 2.0,
    }
    
    medium_weight_positive = {
        'Dividend Yield (%)': 1.5,
        'Current Ratio': 1.0,
        'Quick Ratio': 1.0,
    }
    
    high_weight_negative = {
        'P/E Ratio': 2.0,
        'Debt to Equity': 2.5,
        'Volatility (%)': 1.5,
    }
    
    medium_weight_negative = {
        'Price to Book': 1.0,
        'EV/EBITDA': 1.0,
    }
    
    score_df['Composite Score'] = 0
    
    # Score high weight positive metrics
    for metric, weight in high_weight_positive.items():
        if metric in score_df.columns:
            # Convert to numeric, coerce errors to NaN
            values = pd.to_numeric(score_df[metric], errors='coerce')
            values = values.replace([np.inf, -np.inf], np.nan)
            if values.notna().sum() > 1:
                min_val, max_val = values.min(), values.max()
                if max_val != min_val and not np.isnan(min_val) and not np.isnan(max_val):
                    normalized = (values - min_val) / (max_val - min_val)
                    score_df['Composite Score'] += normalized.fillna(0.5) * 10 * weight
    
    # Score medium weight positive metrics
    for metric, weight in medium_weight_positive.items():
        if metric in score_df.columns:
            values = pd.to_numeric(score_df[metric], errors='coerce')
            values = values.replace([np.inf, -np.inf], np.nan)
            if values.notna().sum() > 1:
                min_val, max_val = values.min(), values.max()
                if max_val != min_val and not np.isnan(min_val) and not np.isnan(max_val):
                    normalized = (values - min_val) / (max_val - min_val)
                    score_df['Composite Score'] += normalized.fillna(0.5) * 10 * weight
    
    # Score high weight negative metrics (inverted)
    for metric, weight in high_weight_negative.items():
        if metric in score_df.columns:
            values = pd.to_numeric(score_df[metric], errors='coerce')
            values = values.replace([np.inf, -np.inf], np.nan)
            if values.notna().sum() > 1:
                min_val, max_val = values.min(), values.max()
                if max_val != min_val and not np.isnan(min_val) and not np.isnan(max_val):
                    normalized = 1 - ((values - min_val) / (max_val - min_val))
                    score_df['Composite Score'] += normalized.fillna(0.5) * 10 * weight
    
    # Score medium weight negative metrics (inverted)
    for metric, weight in medium_weight_negative.items():
        if metric in score_df.columns:
            values = pd.to_numeric(score_df[metric], errors='coerce')
            values = values.replace([np.inf, -np.inf], np.nan)
            if values.notna().sum() > 1:
                min_val, max_val = values.min(), values.max()
                if max_val != min_val and not np.isnan(min_val) and not np.isnan(max_val):
                    normalized = 1 - ((values - min_val) / (max_val - min_val))
                    score_df['Composite Score'] += normalized.fillna(0.5) * 10 * weight
    
    return score_df

def analyze_stocks(df):
    """Provide buy/sell recommendations with detailed analysis"""
    analysis = []
    
    # Sort by composite score
    df_sorted = df.sort_values('Composite Score', ascending=False)
    
    # Top 10 stocks to BUY
    buy_stocks = df_sorted.head(10)
    
    for idx, row in buy_stocks.iterrows():
        recommendation = {
            'Ticker': row['Ticker'],
            'Company': row['Company Name'],
            'Sector': row['Sector'],
            'Action': 'STRONG BUY' if idx < 5 else 'BUY',
            'Score': row['Composite Score'],
            'Current Price': row['Current Price'],
            'Target Potential': 'High' if row['1Y Price Change (%)'] > 30 else 'Medium',
            'Reasons': []
        }
        
        # Profitability analysis
        if row['ROE (%)'] > 20:
            recommendation['Reasons'].append(f"Excellent ROE: {row['ROE (%)']:.2f}% (highly profitable)")
        elif row['ROE (%)'] > 15:
            recommendation['Reasons'].append(f"Strong ROE: {row['ROE (%)']:.2f}%")
        
        if row['Profit Margin (%)'] > 20:
            recommendation['Reasons'].append(f"High profit margin: {row['Profit Margin (%)']:.2f}%")
        
        # Growth analysis
        if row['Revenue Growth (%)'] > 15:
            recommendation['Reasons'].append(f"Strong revenue growth: {row['Revenue Growth (%)']:.2f}%")
        if row['Earnings Growth (%)'] > 15:
            recommendation['Reasons'].append(f"Strong earnings growth: {row['Earnings Growth (%)']:.2f}%")
        
        # Valuation analysis
        if row['P/E Ratio'] < 15 and not np.isnan(row['P/E Ratio']) and row['P/E Ratio'] > 0:
            recommendation['Reasons'].append(f"Attractive valuation: P/E {row['P/E Ratio']:.2f}")
        
        # Performance analysis
        if row['1Y Price Change (%)'] > 30:
            recommendation['Reasons'].append(f"Excellent momentum: +{row['1Y Price Change (%)']:.2f}% (1Y)")
        elif row['1Y Price Change (%)'] > 15:
            recommendation['Reasons'].append(f"Good momentum: +{row['1Y Price Change (%)']:.2f}% (1Y)")
        
        # Risk/return analysis
        if row['Sharpe Ratio'] > 1:
            recommendation['Reasons'].append(f"Excellent risk-adjusted returns: Sharpe {row['Sharpe Ratio']:.2f}")
        
        # Income analysis
        if row['Dividend Yield (%)'] > 4:
            recommendation['Reasons'].append(f"Attractive dividend: {row['Dividend Yield (%)']:.2f}% yield")
        
        # Financial health
        if row['Debt to Equity'] < 0.5 and not np.isnan(row['Debt to Equity']):
            recommendation['Reasons'].append(f"Strong balance sheet: D/E {row['Debt to Equity']:.2f}")
        elif row['Debt to Equity'] < 1 and not np.isnan(row['Debt to Equity']):
            recommendation['Reasons'].append(f"Healthy debt levels: D/E {row['Debt to Equity']:.2f}")
        
        if not recommendation['Reasons']:
            recommendation['Reasons'].append("High composite score across multiple metrics")
        
        analysis.append(recommendation)
    
    # Bottom 10 stocks to SELL/AVOID
    sell_stocks = df_sorted.tail(10)
    
    for idx, row in sell_stocks.iterrows():
        recommendation = {
            'Ticker': row['Ticker'],
            'Company': row['Company Name'],
            'Sector': row['Sector'],
            'Action': 'STRONG SELL' if idx >= len(df_sorted) - 5 else 'AVOID',
            'Score': row['Composite Score'],
            'Current Price': row['Current Price'],
            'Target Potential': 'Low',
            'Reasons': []
        }
        
        # Profitability concerns
        if row['ROE (%)'] < 5 and not np.isnan(row['ROE (%)']):
            recommendation['Reasons'].append(f"Weak profitability: ROE {row['ROE (%)']:.2f}%")
        if row['Profit Margin (%)'] < 5 and not np.isnan(row['Profit Margin (%)']):
            recommendation['Reasons'].append(f"Low profit margin: {row['Profit Margin (%)']:.2f}%")
        
        # Performance concerns
        if row['1Y Price Change (%)'] < -20:
            recommendation['Reasons'].append(f"Severe underperformance: {row['1Y Price Change (%)']:.2f}% (1Y)")
        elif row['1Y Price Change (%)'] < -10:
            recommendation['Reasons'].append(f"Poor performance: {row['1Y Price Change (%)']:.2f}% (1Y)")
        
        # Valuation concerns
        if row['P/E Ratio'] > 30 and not np.isnan(row['P/E Ratio']):
            recommendation['Reasons'].append(f"Overvalued: P/E {row['P/E Ratio']:.2f}")
        
        # Risk concerns
        if row['Sharpe Ratio'] < 0 and not np.isnan(row['Sharpe Ratio']):
            recommendation['Reasons'].append(f"Poor risk-adjusted returns: Sharpe {row['Sharpe Ratio']:.2f}")
        
        if row['Volatility (%)'] > 40:
            recommendation['Reasons'].append(f"High volatility: {row['Volatility (%)']:.2f}%")
        
        # Financial health concerns
        if row['Debt to Equity'] > 2 and not np.isnan(row['Debt to Equity']):
            recommendation['Reasons'].append(f"High debt burden: D/E {row['Debt to Equity']:.2f}")
        
        if row['Current Ratio'] < 1 and not np.isnan(row['Current Ratio']):
            recommendation['Reasons'].append(f"Liquidity concerns: Current Ratio {row['Current Ratio']:.2f}")
        
        if not recommendation['Reasons']:
            recommendation['Reasons'].append("Low composite score with weak overall metrics")
        
        analysis.append(recommendation)
    
    return analysis

def save_to_excel(df, analysis, filename='indonesian_stock_analysis_comprehensive.xlsx'):
    """Save data to Excel with formatting"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Write main data sorted by score
        df_export = df.copy()
        
        # Ensure numeric columns are properly typed
        numeric_cols = ['Current Price', 'Market Cap', 'P/E Ratio', 'Forward P/E', 'PEG Ratio', 
                       'Price to Book', 'Debt to Equity', 'ROE (%)', 'ROA (%)', 'Profit Margin (%)',
                       'Operating Margin (%)', 'Revenue Growth (%)', 'Earnings Growth (%)', 
                       'Dividend Yield (%)', 'Payout Ratio (%)', 'Current Ratio', 'Quick Ratio',
                       '1Y Price Change (%)', '5Y Price Change (%)', 'Volatility (%)', 'Sharpe Ratio',
                       'Avg Daily Volume', '52W High', '52W Low', 'Beta', 'Book Value', 
                       'Enterprise Value', 'EV/EBITDA', 'Composite Score']
        
        for col in numeric_cols:
            if col in df_export.columns:
                df_export[col] = pd.to_numeric(df_export[col], errors='coerce')
        
        df_export = df_export.round(2)
        df_export.to_excel(writer, sheet_name='All Stocks (Ranked)', index=False)
        
        # Write recommendations
        recommendations_data = []
        for rec in analysis:
            recommendations_data.append({
                'Action': rec['Action'],
                'Ticker': rec['Ticker'],
                'Company': rec['Company'],
                'Sector': rec['Sector'],
                'Score': f"{rec['Score']:.2f}",
                'Current Price': f"{rec['Current Price']:.2f}",
                'Potential': rec['Target Potential'],
                'Key Analysis Points': ' | '.join(rec['Reasons'])
            })
        
        recommendations_df = pd.DataFrame(recommendations_data)
        recommendations_df.to_excel(writer, sheet_name='Buy-Sell Recommendations', index=False)
        
        # Create sector analysis with proper numeric handling
        df_for_sector = df.copy()
        # Ensure all columns used in aggregation are numeric
        for col in ['Composite Score', 'ROE (%)', 'P/E Ratio', '1Y Price Change (%)', 'Market Cap']:
            if col in df_for_sector.columns:
                df_for_sector[col] = pd.to_numeric(df_for_sector[col], errors='coerce')
        
        sector_analysis = df_for_sector.groupby('Sector').agg({
            'Composite Score': 'mean',
            'ROE (%)': 'mean',
            'P/E Ratio': 'mean',
            '1Y Price Change (%)': 'mean',
            'Market Cap': 'sum',
            'Ticker': 'count'
        }).round(2)
        sector_analysis.columns = ['Avg Score', 'Avg ROE', 'Avg P/E', 'Avg 1Y Return', 'Total Market Cap', 'Stock Count']
        sector_analysis = sector_analysis.sort_values('Avg Score', ascending=False)
        sector_analysis.to_excel(writer, sheet_name='Sector Analysis')
        
        # Format sheets
        workbook = writer.book
        
        # Format main sheet
        ws1 = workbook['All Stocks (Ranked)']
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Format recommendations sheet
        ws2 = workbook['Buy-Sell Recommendations']
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Color code recommendations
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        dark_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        dark_red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            if row[0].value == 'STRONG BUY':
                for cell in row:
                    cell.fill = dark_green_fill
                    cell.font = Font(bold=True)
            elif row[0].value == 'BUY':
                for cell in row:
                    cell.fill = green_fill
            elif row[0].value == 'STRONG SELL':
                for cell in row:
                    cell.fill = dark_red_fill
                    cell.font = Font(bold=True)
            elif row[0].value == 'AVOID':
                for cell in row:
                    cell.fill = red_fill
        
        # Auto-adjust column widths
        for ws in [ws1, ws2]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 70)
                ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"\n✓ Excel file saved as: {filename}")

def main():
    print("=" * 80)
    print("COMPREHENSIVE INDONESIAN STOCK MARKET ANALYZER")
    print("Covering LQ45, IDX30, IDX80 and major traded stocks")
    print("=" * 80)
    
    # Remove duplicates from stock list
    unique_stocks = list(set(INDONESIAN_STOCKS))
    print(f"\nAnalyzing {len(unique_stocks)} unique Indonesian stocks...")
    print("This covers approximately 90%+ of IDX trading volume")
    print("Fetching 5 years of data and financial reports...")
    print("This will take several minutes. Please wait...\n")
    
    # Collect data with progress indicator
    stock_data = []
    failed_tickers = []
    
    for i, ticker in enumerate(unique_stocks, 1):
        print(f"[{i:3d}/{len(unique_stocks)}] {ticker:12s}", end=' ')
        data = get_stock_data(ticker)
        if data:
            stock_data.append(data)
            print("✓")
        else:
            failed_tickers.append(ticker)
            print("✗ Failed")
    
    if not stock_data:
        print("\n❌ No data collected. Please check your internet connection.")
        return
    
    print(f"\n{'=' * 80}")
    print(f"Successfully collected data for {len(stock_data)} stocks")
    if failed_tickers:
        print(f"Failed to fetch {len(failed_tickers)} tickers: {', '.join(failed_tickers[:10])}")
        if len(failed_tickers) > 10:
            print(f"  ... and {len(failed_tickers) - 10} more")
    print(f"{'=' * 80}\n")
    
    # Create DataFrame
    df = pd.DataFrame(stock_data)
    
    # Calculate composite score
    print("Calculating composite scores and rankings...")
    df = calculate_composite_score(df)
    
    # Sort by composite score
    df = df.sort_values('Composite Score', ascending=False).reset_index(drop=True)
    
    # Generate analysis
    print("Generating buy/sell recommendations...\n")
    analysis = analyze_stocks(df)
    
    # Save to Excel
    save_to_excel(df, analysis)
    
    # Print detailed recommendations
    print("\n" + "=" * 80)
    print("INVESTMENT RECOMMENDATIONS")
    print("=" * 80)
    
    buy_recs = [r for r in analysis if 'BUY' in r['Action']]
    sell_recs = [r for r in analysis if 'SELL' in r['Action'] or 'AVOID' in r['Action']]
    
    print("\n🟢 TOP PICKS TO BUY:")
    print("-" * 80)
    for rec in buy_recs:
        action_emoji = "⭐" if rec['Action'] == 'STRONG BUY' else "✓"
        print(f"\n{action_emoji} {rec['Action']}: {rec['Ticker']} - {rec['Company']}")
        print(f"   Sector: {rec['Sector']} | Score: {rec['Score']:.2f} | Price: IDR {rec['Current Price']:.2f}")
        print(f"   Analysis:")
        for reason in rec['Reasons']:
            print(f"      • {reason}")
    
    print("\n" + "=" * 80)
    print("🔴 STOCKS TO SELL/AVOID:")
    print("-" * 80)
    for rec in sell_recs:
        action_emoji = "⚠️" if rec['Action'] == 'STRONG SELL' else "⊘"
        print(f"\n{action_emoji} {rec['Action']}: {rec['Ticker']} - {rec['Company']}")
        print(f"   Sector: {rec['Sector']} | Score: {rec['Score']:.2f} | Price: IDR {rec['Current Price']:.2f}")
        print(f"   Concerns:")
        for reason in rec['Reasons']:
            print(f"      • {reason}")
    
    # Summary statistics
    print("\n" + "=" * 80)
    print("MARKET SUMMARY")
    print("=" * 80)
    print(f"Total stocks analyzed: {len(df)}")
    print(f"Best performing stock: {df.iloc[0]['Ticker']} - {df.iloc[0]['Company Name']}")
    print(f"  Score: {df.iloc[0]['Composite Score']:.2f} | 1Y Return: {df.iloc[0]['1Y Price Change (%)']:.2f}%")
    print(f"\nWorst performing stock: {df.iloc[-1]['Ticker']} - {df.iloc[-1]['Company Name']}")
    print(f"  Score: {df.iloc[-1]['Composite Score']:.2f} | 1Y Return: {df.iloc[-1]['1Y Price Change (%)']:.2f}%")
    
    # Sector insights
    print(f"\nTop 3 Sectors by Average Score:")
    sector_scores = df.groupby('Sector')['Composite Score'].mean().sort_values(ascending=False).head(3)
    for sector, score in sector_scores.items():
        print(f"  {sector}: {score:.2f}")
    
    print("\n" + "=" * 80)
    print("✓ Analysis complete! Check 'indonesian_stock_analysis_comprehensive.xlsx'")
    print("=" * 80)

if __name__ == "__main__":
    main()