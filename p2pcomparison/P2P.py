#os.chdir(os.path.dirname(os.path.abspath(__file__)))
import pandas as pd
import openpyxl
import os
import glob
os.chdir(os.path.dirname(os.path.abspath(__file__)))

def extract_financial_data(code):
    pattern = f"*{code}*.xlsx"
    files = glob.glob(pattern)
    if not files:
        print(f"Warning: No file found for code {code}")
        return {
            'Industry': '-1',
            'Subindustry': '-1',
            'Currency': '-1',
            'Unit': '-1',
            'Asset': -1,
            'Liability': -1,
            'Equity': -1,
            'Revenue': -1,
            'Gross Profit': -1,
            'Net Profit': -1
        }
    file_path = files[0]
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        data = {
            'Industry': '-1',
            'Subindustry': '-1',
            'Currency': '-1',
            'Unit': '-1',
            'Asset': -1,
            'Liability': -1,
            'Equity': -1,
            'Revenue': -1,
            'Gross Profit': -1,
            'Net Profit': -1
        }
        if '1000000' in wb.sheetnames:
            if wb['1000000']['A14'].value == 'Industri':
                data['Industry'] = wb['1000000']['B14'].value
            if wb['1000000']['A15'].value == 'Subindustri':
                data['Subindustry'] = wb['1000000']['B15'].value
            if wb['1000000']['A29'].value == 'Mata uang pelaporan':
                data['Currency'] = wb['1000000']['B29'].value
            if wb['1000000']['A31'].value == 'Pembulatan yang digunakan dalam penyajian jumlah dalam laporan keuangan':
                data['Unit'] = wb['1000000']['B31'].value
        if '1210000' in wb.sheetnames:
            if wb['1210000']['A128'].value == "Jumlah aset":
                data['Asset'] = wb['1210000']['B128'].value
            if wb['1210000']['A247'].value == "Jumlah liabilitas":
                data['Liability'] = wb['1210000']['B247'].value
            if wb['1210000']['A272'].value == "Jumlah ekuitas":
                data['Equity'] = wb['1210000']['B272'].value
        if '1321000' in wb.sheetnames:
            if wb['1321000']['A6'].value == "Penjualan dan pendapatan usaha":
                data['Revenue'] = wb['1321000']['B6'].value
            if wb['1321000']['A8'].value == "Jumlah laba bruto":
                data['Gross Profit'] = wb['1321000']['B8'].value
            if wb['1321000']['A30'].value == "Jumlah laba (rugi)":
                data['Net Profit'] = wb['1321000']['B30'].value
        elif '1311000' in wb.sheetnames:
            if wb['1311000']['A6'].value == "Penjualan dan pendapatan usaha":
                data['Revenue'] = wb['1311000']['B6'].value
            if wb['1311000']['A8'].value == "Jumlah laba bruto":
                data['Gross Profit'] = wb['1311000']['B8'].value
            if wb['1311000']['A30'].value == "Jumlah laba (rugi)":
                data['Net Profit'] = wb['1311000']['B30'].value
        wb.close()
        return data
        
    except Exception as e:
        print(f"Error processing {code}: {str(e)}")
        return {
            'Industry': '-1',
            'Subindustry': '-1',
            'Currency': '-1',
            'Unit': '-1',
            'Asset': -1,
            'Liability': -1,
            'Equity': -1,
            'Revenue': -1,
            'Gross Profit': -1,
            'Net Profit': -1
        }

def calculate_ratios(row):
    ratios = {}
    ratios['D/A'] = row['Liability'] / row['Asset']
    ratios['E/A'] = row['Equity'] / row['Asset']
    ratios['ROE'] = row['Net Profit'] / row['Equity']
    ratios['ROA'] = row['Net Profit'] / row['Asset']
    ratios['Gross Margin'] = row['Gross Profit'] / row['Revenue']
    ratios['Net Margin'] = row['Net Profit'] / row['Revenue']
    return ratios

def main():
    input_file = "Daftar Saham  - Healthcare - 20251014.xlsx"
    df = pd.read_excel(input_file)
    print(f"Loaded {len(df)} stocks from {input_file}")
    financial_columns = ['Industry', 'Subindustry', 'Currency', 'Unit', 'Asset', 'Liability', 'Equity', 'Revenue', 'Gross Profit', 'Net Profit']
    for col in financial_columns:
        df[col] = 0
    print("\nExtracting financial data...")
    for idx, row in df.iterrows():
        code = row['Kode']
        print(f"Processing {code}...")
        financial_data = extract_financial_data(code)
        for col in financial_columns:
            df.at[idx, col] = financial_data[col]
    
    print("\nCalculating financial ratios...")
    ratio_columns = ['D/A', 'E/A', 'ROE', 'ROA', 'Gross Margin', 'Net Margin']
    for col in ratio_columns:
        df[col] = 0
    
    for idx, row in df.iterrows():
        ratios = calculate_ratios(row)
        for ratio_name, ratio_value in ratios.items():
            df.at[idx, ratio_name] = ratios[ratio_name]
    
    df.to_excel('result.xlsx', index=False)
    print("\nResults saved to result.xlsx")
    print(f"Total stocks processed: {len(df)}")

if __name__ == "__main__":
    main()